import json
import csv
import io
from typing import List
from ..models import JiraExportInput, JiraExportResponse, TestCase


def export_to_jira(payload: JiraExportInput) -> JiraExportResponse:
    """Export test cases to JIRA (stub - requires JIRA credentials)."""
    return JiraExportResponse(
        status="stubbed",
        message=(
            f"JIRA export adapter not configured. Would export {len(payload.test_cases)} "
            f"test cases to project '{payload.project_key}' as '{payload.issue_type}' issues. "
            "Provide JIRA credentials and API configuration to enable this feature."
        ),
    )


def export_to_csv(test_cases: List[TestCase]) -> str:
    """Export test cases to CSV format."""
    output = io.StringIO()
    
    # Define headers based on industry standard fields
    headers = [
        "ID", "Title", "Description", "Priority", "Type", "Status",
        "Preconditions", "Steps", "Expected Result", "Test Data",
        "Estimated Time", "Automation Status", "Component", "Tags"
    ]
    
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow(headers)
    
    for tc in test_cases:
        # Format steps as numbered list
        steps_text = "\n".join([
            f"{s.step}. {s.action} -> Expected: {s.expected}" + 
            (f" [Data: {s.test_data}]" if s.test_data else "")
            for s in tc.steps
        ])
        
        row = [
            tc.id,
            tc.title,
            tc.description or "",
            tc.priority,
            tc.type,
            tc.status,
            tc.preconditions or "",
            steps_text,
            tc.expected_result or "",
            tc.test_data or "",
            tc.estimated_time or "",
            tc.automation_status,
            tc.component or "",
            ", ".join(tc.tags) if tc.tags else ""
        ]
        writer.writerow(row)
    
    return output.getvalue()


def export_to_excel(test_cases: List[TestCase]) -> bytes:
    """Export test cases to Excel format (XLSX)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        # Fallback to CSV-compatible format if openpyxl not installed
        csv_content = export_to_csv(test_cases)
        return csv_content.encode('utf-8')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Define headers
    headers = [
        "ID", "Title", "Description", "Priority", "Type", "Status",
        "Preconditions", "Steps", "Expected Result", "Test Data",
        "Estimated Time", "Automation Status", "Component", "Tags"
    ]
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Priority colors
    priority_colors = {
        "Critical": "FF6B6B",
        "High": "FFA500",
        "Medium": "FFD93D",
        "Low": "6BCB77"
    }
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row_idx, tc in enumerate(test_cases, 2):
        # Format steps
        steps_text = "\n".join([
            f"{s.step}. {s.action}\n   → {s.expected}" + 
            (f"\n   [Data: {s.test_data}]" if s.test_data else "")
            for s in tc.steps
        ])
        
        data = [
            tc.id,
            tc.title,
            tc.description or "",
            tc.priority,
            tc.type,
            tc.status,
            tc.preconditions or "",
            steps_text,
            tc.expected_result or "",
            tc.test_data or "",
            tc.estimated_time or "",
            tc.automation_status,
            tc.component or "",
            ", ".join(tc.tags) if tc.tags else ""
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Color-code priorities
            if col == 4 and value in priority_colors:
                cell.fill = PatternFill(start_color=priority_colors[value], 
                                        end_color=priority_colors[value], 
                                        fill_type="solid")
    
    # Adjust column widths
    column_widths = [12, 40, 50, 10, 15, 12, 40, 60, 40, 30, 15, 18, 20, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_to_json(test_cases: List[TestCase]) -> str:
    """Export test cases to JSON format."""
    # Convert to dict with proper serialization
    data = {
        "export_format": "test_cases_v1",
        "total_count": len(test_cases),
        "test_cases": [
            {
                "id": tc.id,
                "title": tc.title,
                "description": tc.description,
                "priority": tc.priority,
                "type": tc.type,
                "status": tc.status,
                "preconditions": tc.preconditions,
                "steps": [
                    {
                        "step": s.step,
                        "action": s.action,
                        "expected": s.expected,
                        "test_data": s.test_data
                    }
                    for s in tc.steps
                ],
                "expected_result": tc.expected_result,
                "test_data": tc.test_data,
                "estimated_time": tc.estimated_time,
                "automation_status": tc.automation_status,
                "component": tc.component,
                "tags": tc.tags
            }
            for tc in test_cases
        ]
    }
    return json.dumps(data, indent=2)
