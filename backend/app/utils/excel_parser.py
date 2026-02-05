"""
Excel parser utility for extracting requirements text from .xlsx files.

Supports two formats with auto-detection:
1. Structured: Each row represents a requirement (with columns like ID, Description, etc.)
2. Free-form: Requirements scattered as text across cells

All sheets in the workbook are processed.
"""

from io import BytesIO
from typing import List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# Keywords that indicate a structured requirements column
REQUIREMENT_KEYWORDS = [
    "requirement",
    "description", 
    "story",
    "user story",
    "feature",
    "acceptance criteria",
    "spec",
    "specification",
]

# Keywords that indicate an ID column
ID_KEYWORDS = ["id", "req id", "requirement id", "story id", "key", "#"]


def _find_header_column(headers: Tuple, keywords: List[str]) -> Optional[int]:
    """Find the column index that matches any of the keywords."""
    for idx, header in enumerate(headers):
        if header is None:
            continue
        header_lower = str(header).lower().strip()
        for keyword in keywords:
            if keyword in header_lower:
                return idx
    return None


def _is_structured_sheet(first_row: Tuple) -> bool:
    """Determine if the first row looks like a header row for structured data."""
    if not first_row or all(cell is None for cell in first_row):
        return False
    
    # Check if any cell matches requirement-related keywords
    for cell in first_row:
        if cell is None:
            continue
        cell_lower = str(cell).lower().strip()
        for keyword in REQUIREMENT_KEYWORDS + ID_KEYWORDS:
            if keyword in cell_lower:
                return True
    return False


def _extract_structured(ws: Worksheet, headers: Tuple) -> List[str]:
    """Extract requirements from a structured sheet (one requirement per row)."""
    # Find the requirement/description column
    req_col = _find_header_column(headers, REQUIREMENT_KEYWORDS)
    id_col = _find_header_column(headers, ID_KEYWORDS)
    
    # If no requirement column found, use the second column (first is often ID)
    if req_col is None:
        req_col = 1 if len(headers) > 1 else 0
    
    requirements = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header row
    
    for row in rows:
        if not row or len(row) <= req_col:
            continue
        
        req_text = row[req_col]
        if req_text is None or str(req_text).strip() == "":
            continue
        
        # Include ID prefix if available
        if id_col is not None and len(row) > id_col and row[id_col]:
            requirement = f"{row[id_col]}: {req_text}"
        else:
            requirement = str(req_text)
        
        requirements.append(requirement.strip())
    
    return requirements


def _extract_freeform(ws: Worksheet) -> List[str]:
    """Extract text from a free-form sheet by concatenating non-empty cells."""
    texts = []
    
    for row in ws.iter_rows(values_only=True):
        row_parts = []
        for cell in row:
            if cell is not None and str(cell).strip():
                row_parts.append(str(cell).strip())
        
        if row_parts:
            row_text = " ".join(row_parts)
            texts.append(row_text)
    
    return texts


def _process_sheet(ws: Worksheet) -> List[str]:
    """Process a single worksheet and extract requirements text."""
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows:
        return []
    
    first_row = rows[0] if rows else tuple()
    
    if _is_structured_sheet(first_row):
        return _extract_structured(ws, first_row)
    else:
        return _extract_freeform(ws)


def parse_excel_to_text(content: bytes) -> str:
    """
    Parse an Excel file and extract requirements as text.
    
    Args:
        content: Raw bytes of the Excel file
        
    Returns:
        Extracted text suitable for requirements parsing
    """
    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    
    all_requirements = []
    
    for ws in wb.worksheets:
        sheet_requirements = _process_sheet(ws)
        if sheet_requirements:
            # Add sheet name as context if multiple sheets have content
            if len(wb.worksheets) > 1 and sheet_requirements:
                all_requirements.append(f"--- {ws.title} ---")
            all_requirements.extend(sheet_requirements)
    
    wb.close()
    
    return "\n".join(all_requirements)
